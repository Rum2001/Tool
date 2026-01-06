import streamlit as st
import pandas as pd
import mysql.connector
import os
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
import qrcode
import qrcode.image.svg
from PIL import Image
import io
import tempfile
import time
import json
import math
import zipfile

def load_csv_connections(uploaded_file):
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file)
            return df.to_dict('records')
        except Exception as e:
            st.error(f'Error reading CSV file: {str(e)}')
    return []

def load_company_queries(uploaded_file):
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file)
            return dict(zip(df['company'], df['query']))
        except Exception as e:
            st.error(f'Error reading company queries CSV file: {str(e)}')
    return {}

def connect_to_database(host, user, password, database):
    try:
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database=database,
            connect_timeout=30,  # Thêm timeout cho Streamlit Cloud
            autocommit=True
        )
        return conn
    except mysql.connector.Error as err:
        st.error(f"Database connection error: {err}")
        return None

def execute_query(conn, query):
    if not conn:
        st.warning('Please connect to a database first.')
        return None
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query)
        results = cursor.fetchall()
        cursor.close()
        return results
    except mysql.connector.Error as err:
        st.error(f"Error executing query: {err}")
        return None

def save_results_to_json(results, filename=None):
    """Save query results to a JSON file for faster processing"""
    try:
        # Sử dụng tempfile để tránh vấn đề permissions trên cloud
        if filename is None:
            temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json', encoding='utf-8')
            filename = temp_file.name
            temp_file.close()
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False)
        return filename
    except Exception as e:
        st.error(f"Error saving results to JSON: {str(e)}")
        return None

def load_results_from_json(filename):
    """Load results from JSON file"""
    try:
        if not filename or not os.path.exists(filename):
            return None
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        st.error(f"Error loading results from JSON: {str(e)}")
        return None

def generate_insert_query_batched(table_name, selected_columns, results, batch_size=1000):
    """Generate INSERT queries in batches"""
    if not results or not selected_columns or not table_name:
        return None
    
    # Calculate total batches
    total_records = len(results)
    total_batches = math.ceil(total_records / batch_size)
    
    # Create progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Sử dụng tempfile cho Streamlit Cloud
    temp_json = None
    temp_sql = None
    
    try:
        # Save results to temp JSON file
        temp_json = save_results_to_json(results)
        if not temp_json:
            return None
        
        # Generate base query
        columns_str = ', '.join(selected_columns)
        base_query = f"INSERT INTO {table_name} ({columns_str}) VALUES\n"
        
        # Tạo temp file cho SQL queries
        temp_sql_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.sql', encoding='utf-8')
        temp_sql = temp_sql_file.name
        
        # Load results from JSON
        results = load_results_from_json(temp_json)
        if not results:
            return None
        
        # Process in batches
        for batch_num in range(total_batches):
            start_idx = batch_num * batch_size
            end_idx = min((batch_num + 1) * batch_size, total_records)
            batch = results[start_idx:end_idx]
            
            values_list = []
            for row in batch:
                values = []
                for col in selected_columns:
                    val = row[col]
                    if val is None:
                        values.append('NULL')
                    elif isinstance(val, (int, float)):
                        values.append(str(val))
                    else:
                        val = str(val).replace("'", "''").replace('"', '\\"')
                        values.append(f"'{val}'")
                values_list.append(f"({', '.join(values)})")
            
            # Write batch query to file
            batch_query = base_query + ',\n'.join(values_list) + ";"
            temp_sql_file.write(batch_query + "\n-- Next batch --\n")
            
            # Update progress
            progress = (batch_num + 1) / total_batches
            progress_bar.progress(progress)
            status_text.text(f"Processing batch {batch_num + 1} of {total_batches}")
        
        temp_sql_file.close()
        
        # Read generated queries
        with open(temp_sql, 'r', encoding='utf-8') as f:
            queries = f.read()
        
        return queries                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
    
    except Exception as e:
        st.error(f"Error generating INSERT queries: {str(e)}")
        return None
    finally:
        # Clean up temp files
        try:
            if temp_json and os.path.exists(temp_json):
                os.remove(temp_json)
            if temp_sql and os.path.exists(temp_sql):
                os.remove(temp_sql)
        except:
            pass

def generate_delete_query(table_name, column, results):
    if not results or not column or not table_name:
        return None
    
    values = [f'"{str(row[column])}"' for row in results]
    values_str = ', '.join(values)
    return f"DELETE FROM {table_name} WHERE {column} IN ({values_str});"

def execute_insert_delete_query(conn, query):
    if not conn:
        st.warning('Please connect to a database first.')
        return False
    
    try:
        cursor = conn.cursor()
        cursor.execute(query)
        conn.commit()
        affected_rows = cursor.rowcount
        cursor.close()
        return affected_rows
    except mysql.connector.Error as err:
        st.error(f"Error executing query: {err}")
        return False

def main():
    st.set_page_config(page_title="Export Code", layout="wide")
    st.title("Export Code")
    
    # Hiển thị warning nếu đang chạy trên Streamlit Cloud
    if os.getenv('STREAMLIT_CLOUD'):
        st.info("🌐 Đang chạy trên Streamlit Cloud. Lưu ý: có giới hạn về RAM và processing time.")
        st.warning("⚠️ Khuyến nghị: Xử lý dữ liệu nhỏ hơn 50MB và dưới 10,000 rows để tránh timeout.")

    # Initialize session state
    if 'connections' not in st.session_state:
        st.session_state['connections'] = []
    if 'company_queries' not in st.session_state:
        st.session_state['company_queries'] = {}
    if 'db_connection' not in st.session_state:
        st.session_state['db_connection'] = None
    if 'query_results' not in st.session_state:
        st.session_state['query_results'] = None

    # Create tabs for different functionalities
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["Database Connection", "Query Execution", "Insert/Delete", "Export Options", "Batch Export", "Excel Upload & QR", "Tra cứu mã lỗi"])

    with tab1:
        st.subheader("Database Connection")
        
        # CSV Upload for connections
        connections_file = st.file_uploader("Upload Connections CSV", type=['csv'])
        if connections_file:
            connections = load_csv_connections(connections_file)
            if connections:
                st.session_state['connections'] = connections
                st.success("Connections CSV loaded successfully!")

        # Connection selection
        if st.session_state['connections']:
            connection_names = [f"{conn['host']} - {conn['database']}" for conn in st.session_state['connections']]
            selected_conn = st.selectbox("Select Connection", ['Select a connection'] + connection_names)
            
            if selected_conn != 'Select a connection':
                idx = connection_names.index(selected_conn)
                conn_details = st.session_state['connections'][idx]
                
                col1, col2 = st.columns(2)
                with col1:
                    host = st.text_input("Host", value=conn_details['host'])
                    user = st.text_input("User", value=conn_details['user'])
                with col2:
                    database = st.text_input("Database", value=conn_details['database'])
                    password = st.text_input("Password", value=conn_details['password'], type="password")

                if st.button("Connect to Database"):
                    st.session_state['db_connection'] = connect_to_database(host, user, password, database)
                    if st.session_state['db_connection']:
                        st.success("Connected to database successfully!")

    with tab2:
        st.subheader("Query Execution")
        
        # Company queries upload
        company_queries_file = st.file_uploader("Upload Company Queries CSV", type=['csv'])
        if company_queries_file:
            company_queries = load_company_queries(company_queries_file)
            if company_queries:
                st.session_state['company_queries'] = company_queries
                st.success("Company queries loaded successfully!")

        # Company selection and query input
        if st.session_state['company_queries']:
            selected_company = st.selectbox("Select Company", ['Select a company'] + list(st.session_state['company_queries'].keys()))
            if selected_company != 'Select a company':
                query = st.text_area("SQL Query", value=st.session_state['company_queries'][selected_company], height=150)
            else:
                query = st.text_area("SQL Query", height=150)
        else:
            query = st.text_area("SQL Query", height=150)

        if st.button("Execute Query"):
            if st.session_state['db_connection'] and query:
                results = execute_query(st.session_state['db_connection'], query)
                if results:
                    st.session_state['query_results'] = results
                    st.dataframe(pd.DataFrame(results))
                    st.success(f"Query executed successfully! {len(results)} rows returned.")
                else:
                    st.warning("No results returned from query.")

    with tab3:
        st.subheader("Insert/Delete Query Generation")
        
        if st.session_state['query_results']:
            # Target database selection
            if st.session_state['connections']:
                target_db = st.selectbox("Target Database for Insert/Delete", 
                                       ['Select target database'] + connection_names,
                                       key='target_db')
            
            # Target table input
            target_table = st.text_input("Target Table Name")
            
            # Column selection for INSERT/DELETE
            if st.session_state['query_results']:
                columns = list(st.session_state['query_results'][0].keys())
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.subheader("INSERT Query")
                    selected_columns = st.multiselect("Select Columns for INSERT", columns)
                    batch_size = st.number_input("Batch Size", min_value=100, value=1000, step=100)
                    
                    if st.button("Generate INSERT Query"):
                        if target_table and selected_columns:
                            st.info("Generating INSERT queries... This may take a while for large datasets.")
                            insert_queries = generate_insert_query_batched(
                                target_table, 
                                selected_columns, 
                                st.session_state['query_results'],
                                batch_size=batch_size
                            )
                            if insert_queries:
                                st.download_button(
                                    "Download INSERT Queries",
                                    insert_queries,
                                    file_name="insert_queries.sql",
                                    mime="text/plain"
                                )
                                st.text_area("Preview of Generated INSERT Queries", 
                                           insert_queries[:1000] + "...", 
                                           height=200)
                
                with col2:
                    st.subheader("DELETE Query")
                    delete_column = st.selectbox("Select Column for DELETE condition", columns)
                    
                    if st.button("Generate DELETE Query"):
                        if target_table and delete_column:
                            delete_query = generate_delete_query(target_table, delete_column, 
                                                              st.session_state['query_results'])
                            if delete_query:
                                st.text_area("Generated DELETE Query", delete_query, height=200)
                                st.download_button(
                                    "Tải xuống câu lệnh DELETE",
                                    delete_query,
                                    file_name=f"{target_table}_delete_query.sql",
                                    mime="text/plain"
                                )
                                if st.button("Execute DELETE"):
                                    if target_db != 'Select target database':
                                        idx = connection_names.index(target_db)
                                        target_conn_details = st.session_state['connections'][idx]
                                        target_conn = connect_to_database(
                                            target_conn_details['host'],
                                            target_conn_details['user'],
                                            target_conn_details['password'],
                                            target_conn_details['database']
                                        )
                                        if target_conn:
                                            affected_rows = execute_insert_delete_query(target_conn, delete_query)
                                            if affected_rows:
                                                st.success(f"Successfully deleted {affected_rows} rows!")
                                            target_conn.close()

    with tab4:
        st.subheader("Export Options")
        
        if st.session_state['query_results']:
            col1, col2 = st.columns(2)
            
            with col1:
                file_prefix = st.text_input("File Prefix", "export")
                rows_per_file = st.number_input("Rows per File", min_value=1, value=9000)
                double_row = st.checkbox("Export double rows")
                include_headers = st.checkbox("Include column headers", value=True)

            with col2:
                if st.session_state['query_results']:
                    columns = list(st.session_state['query_results'][0].keys())
                    qr_column = st.selectbox("QR Code Column", columns)
                    image_name_column = st.selectbox("Image Name Column", columns)
                    qr_format = st.selectbox("QR Code Format", ["png", "jpg", "svg"])

            # Export buttons
            col3, col4 = st.columns(2)
            with col3:
                if st.button("Export to Excel"):
                    try:
                        temp_dir = tempfile.mkdtemp()
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        total_rows = len(st.session_state['query_results'])
                        chunks = [st.session_state['query_results'][i:i + rows_per_file] 
                                for i in range(0, total_rows, rows_per_file)]
                        
                        # Create a zip file to store all Excel files
                        zip_path = os.path.join(temp_dir, f"{file_prefix}_excel_files.zip")
                        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                            for i, chunk in enumerate(chunks):
                                file_name = f'{file_prefix}-{i+1:03d}.xlsx'
                                file_path = os.path.join(temp_dir, file_name)
                                
                                wb = openpyxl.Workbook()
                                ws = wb.active
                                
                                headers = list(chunk[0].keys())
                                row_num = 1
                                
                                # Write headers only if include_headers is True
                                if include_headers:
                                    for col, header in enumerate(headers, 1):
                                        cell = ws.cell(row=1, column=col, value=header)
                                        cell.font = Font(bold=True)
                                        cell.alignment = Alignment(horizontal='center')
                                    row_num = 2
                                
                                # Write data
                                for record in chunk:
                                    for col, key in enumerate(headers, 1):
                                        ws.cell(row=row_num, column=col, value=str(record[key]))
                                    if double_row:
                                        row_num += 1
                                        for col, key in enumerate(headers, 1):
                                            ws.cell(row=row_num, column=col, value=str(record[key]))
                                    row_num += 1
                                
                                wb.save(file_path)
                                
                                # Add Excel file to zip
                                zipf.write(file_path, file_name)
                                
                                progress = (i + 1) / len(chunks)
                                progress_bar.progress(progress)
                                status_text.text(f"Processing file {i+1} of {len(chunks)}")
                        
                        # Create download button for zip file
                        with open(zip_path, 'rb') as f:
                            zip_contents = f.read()
                            st.download_button(
                                "Download All Excel Files (ZIP)",
                                zip_contents,
                                file_name=f"{file_prefix}_excel_files.zip",
                                mime="application/zip"
                            )
                        
                        st.success("Excel export completed successfully!")
                    except Exception as e:
                        st.error(f"Error during export: {str(e)}")
                    finally:
                        # Clean up temporary directory
                        import shutil
                        shutil.rmtree(temp_dir, ignore_errors=True)

            with col4:
                col4_1, col4_2, col4_3 = st.columns(3)
                
                with col4_1:
                    # Add table name input
                    sql_table_name = st.text_input("SQL Table Name", key="sql_table_name")
                    
                    if st.button("Export to SQL"):
                        if not sql_table_name:
                            st.error("Please enter a table name")
                        else:
                            try:
                                temp_dir = tempfile.mkdtemp()
                                progress_bar = st.progress(0)
                                status_text = st.empty()
                                
                                total_rows = len(st.session_state['query_results'])
                                chunks = [st.session_state['query_results'][i:i + rows_per_file] 
                                        for i in range(0, total_rows, rows_per_file)]
                                
                                # Create a zip file to store all SQL files
                                zip_path = os.path.join(temp_dir, f"{file_prefix}_sql_files.zip")
                                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                                    for i, chunk in enumerate(chunks):
                                        if not chunk:
                                            continue
                                            
                                        file_name = f'{file_prefix}-{i+1:03d}.sql'
                                        file_path = os.path.join(temp_dir, file_name)
                                        
                                        try:
                                            # Get column names from first row
                                            columns = list(chunk[0].keys())
                                            columns_str = ', '.join(f"`{col}`" for col in columns)
                                            
                                            # Write INSERT statements
                                            values_list = []
                                            for record in chunk:
                                                values = []
                                                for col in columns:
                                                    val = record[col]
                                                    if val is None:
                                                        values.append('NULL')
                                                    elif isinstance(val, (int, float)):
                                                        values.append(str(val))
                                                    else:
                                                        # Properly escape string values
                                                        val = str(val).replace('\\', '\\\\').replace("'", "\\'").replace('"', '\\"')
                                                        values.append(f"'{val}'")
                                                values_list.append(f"({', '.join(values)})")
                                            
                                            # Write all values in one INSERT statement with proper formatting
                                            insert_query = f"INSERT INTO `{sql_table_name}` ({columns_str}) VALUES\n"
                                            insert_query += ',\n'.join(values_list) + ";\n"
                                            
                                            # Write to file
                                            with open(file_path, 'w', encoding='utf-8') as f:
                                                f.write(insert_query)
                                                if double_row:
                                                    f.write(insert_query)
                                            
                                            # Add the SQL file to the zip
                                            zipf.write(file_path, file_name)
                                            
                                        except Exception as e:
                                            st.error(f"Error generating SQL for chunk {i+1}: {str(e)}")
                                            continue
                                        
                                        # Update progress
                                        progress = (i + 1) / len(chunks)
                                        progress_bar.progress(progress)
                                        status_text.text(f"Processing file {i+1} of {len(chunks)}")
                                
                                # Create download button for zip file
                                with open(zip_path, 'rb') as f:
                                    zip_contents = f.read()
                                    st.download_button(
                                        "Download All SQL Files (ZIP)",
                                        zip_contents,
                                        file_name=f"{file_prefix}_sql_files.zip",
                                        mime="application/zip"
                                    )
                                
                                st.success("SQL export completed successfully!")
                            except Exception as e:
                                st.error(f"Error during SQL export: {str(e)}")
                            finally:
                                # Clean up temporary directory
                                import shutil
                                shutil.rmtree(temp_dir, ignore_errors=True)
                
                with col4_2:
                    if st.button("Export to TXT"):
                        try:
                            temp_dir = tempfile.mkdtemp()
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            total_rows = len(st.session_state['query_results'])
                            chunks = [st.session_state['query_results'][i:i + rows_per_file] 
                                    for i in range(0, total_rows, rows_per_file)]
                            
                            # Create a zip file to store all TXT files
                            zip_path = os.path.join(temp_dir, f"{file_prefix}_txt_files.zip")
                            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                                for i, chunk in enumerate(chunks):
                                    file_name = f'{file_prefix}-{i+1:03d}.txt'
                                    file_path = os.path.join(temp_dir, file_name)
                                    
                                    with open(file_path, 'w', encoding='utf-8', newline='') as txtfile:
                                        headers = list(chunk[0].keys())
                                        
                                        # Write headers if include_headers is True
                                        if include_headers:
                                            txtfile.write(','.join(headers) + '\n')
                                        
                                        # Write data with actual values separated by commas
                                        for record in chunk:
                                            row_data = []
                                            for key in headers:
                                                value = record[key]
                                                if value is None:
                                                    row_data.append('')
                                                else:
                                                    # Convert to string and handle special characters
                                                    str_value = str(value).replace('\n', ' ').replace('\r', ' ')
                                                    # Escape commas in data by wrapping in quotes
                                                    if ',' in str_value:
                                                        str_value = f'"{str_value}"'
                                                    row_data.append(str_value)
                                            
                                            data_line = ','.join(row_data)
                                            txtfile.write(data_line + '\n')
                                            
                                            # Write double row if option is selected
                                            if double_row:
                                                txtfile.write(data_line + '\n')
                                    
                                    # Add TXT file to zip
                                    zipf.write(file_path, file_name)
                                    
                                    progress = (i + 1) / len(chunks)
                                    progress_bar.progress(progress)
                                    status_text.text(f"Processing file {i+1} of {len(chunks)}")
                            
                            # Create download button for zip file
                            with open(zip_path, 'rb') as f:
                                zip_contents = f.read()
                                st.download_button(
                                    "Download All TXT Files (ZIP)",
                                    zip_contents,
                                    file_name=f"{file_prefix}_txt_files.zip",
                                    mime="application/zip"
                                )
                            
                            st.success("TXT export completed successfully!")
                        except Exception as e:
                            st.error(f"Error during TXT export: {str(e)}")
                        finally:
                            # Clean up temporary directory
                            import shutil
                            shutil.rmtree(temp_dir, ignore_errors=True)
                
                with col4_3:
                    if st.button("Export QR Codes"):
                        try:
                            temp_dir = tempfile.mkdtemp()
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            # Create a zip file for QR codes
                            zip_path = os.path.join(temp_dir, f"{file_prefix}_qr_codes.zip")
                            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                                total = len(st.session_state['query_results'])
                                for i, record in enumerate(st.session_state['query_results']):
                                    qr_data = str(record[qr_column])
                                    
                                    # Clean filename - remove invalid characters
                                    raw_name = str(record[image_name_column])
                                    # Replace newlines, tabs, and other control characters with space
                                    clean_name = raw_name.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                                    # Remove invalid Windows filename characters
                                    invalid_chars = '<>:"/\\|?*'
                                    for char in invalid_chars:
                                        clean_name = clean_name.replace(char, '_')
                                    # Remove leading/trailing spaces and dots
                                    clean_name = clean_name.strip('. ')
                                    # Limit filename length (Windows has 255 char limit)
                                    if len(clean_name) > 200:
                                        clean_name = clean_name[:200]
                                    # Ensure filename is not empty
                                    if not clean_name:
                                        clean_name = f"qr_code_{i+1}"
                                    
                                    image_name = f"{clean_name}.{qr_format}"
                                    file_path = os.path.join(temp_dir, image_name)
                                    
                                    qr = qrcode.QRCode(version=1, box_size=10, border=5)
                                    qr.add_data(qr_data)
                                    qr.make(fit=True)
                                    
                                    # Save based on format
                                    if qr_format == 'svg':
                                        # For SVG, use the make_image with factory
                                        factory = qrcode.image.svg.SvgPathImage
                                        qr_svg = qrcode.QRCode(
                                            version=1,
                                            box_size=10,
                                            border=5,
                                            image_factory=factory
                                        )
                                        qr_svg.add_data(qr_data)
                                        qr_svg.make(fit=True)
                                        img_svg = qr_svg.make_image(fill_color="black", back_color="white")
                                        img_svg.save(file_path)
                                    elif qr_format == 'png':
                                        img = qr.make_image(fill_color="black", back_color="white")
                                        img.save(file_path, format='PNG')
                                    elif qr_format == 'jpg':
                                        # Convert to RGB mode for JPEG (no transparency)
                                        img = qr.make_image(fill_color="black", back_color="white")
                                        img_rgb = img.convert('RGB')
                                        img_rgb.save(file_path, format='JPEG', quality=95)
                                    
                                    # Add QR code to zip
                                    zipf.write(file_path, image_name)
                                    
                                    progress = (i + 1) / total
                                    progress_bar.progress(progress)
                                    status_text.text(f"Processing QR code {i+1} of {total}")
                            
                            # Create download button for zip file
                            with open(zip_path, 'rb') as f:
                                zip_contents = f.read()
                                st.download_button(
                                    "Download QR Codes (ZIP)",
                                    zip_contents,
                                    file_name=f"{file_prefix}_qr_codes.zip",
                                    mime="application/zip"
                                )
                            
                            st.success("QR codes exported successfully!")
                        except Exception as e:
                            st.error(f"Error during QR code export: {str(e)}")
                        finally:
                            # Clean up temporary directory
                            import shutil
                            shutil.rmtree(temp_dir, ignore_errors=True)

    # Initialize batch results in session state
    if 'batch_results' not in st.session_state:
        st.session_state['batch_results'] = None
    
    with tab5:
        st.subheader("Batch Export - Xuất nhiều lệnh SQL cùng lúc")
        
        st.info("💡 Nhập các câu lệnh SQL (mỗi dòng một lệnh), sau đó xem kết quả và đặt tên file cho mỗi bảng")
        
        # Step 1: Input queries and execute
        st.markdown("### Bước 1: Nhập và Thực thi các lệnh SQL")
        
        # Text area for multiple queries
        batch_queries_input = st.text_area(
            "Nhập các lệnh SQL (mỗi dòng một lệnh)",
            height=200,
            placeholder="Ví dụ:\nSELECT id, CONCAT('http://sh.vinachg.vn/ck/?s=', `serial_rand`) AS `qrcode`, `serial` FROM stamp_sh WHERE stamp_block_id = 2870 ORDER BY `serial` ASC\nSELECT id, CONCAT('http://sh.vinachg.vn/ck/?s=', `serial_rand`) AS `qrcode`, `serial` FROM stamp_sh WHERE stamp_block_id = 2871 ORDER BY `serial` ASC"
        )
        
        col_exec1, col_exec2 = st.columns([4, 1])
        
        with col_exec1:
            if st.button("🔍 Thực thi và Xem Kết quả", key="batch_execute", type="primary"):
                if not st.session_state['db_connection']:
                    st.error("❌ Vui lòng kết nối database trước!")
                elif not batch_queries_input.strip():
                    st.error("❌ Vui lòng nhập ít nhất một lệnh SQL!")
                else:
                    try:
                        # Parse input
                        lines = batch_queries_input.strip().split('\n')
                        queries_list = []
                        
                        for idx, line in enumerate(lines):
                            line = line.strip()
                            if not line or line.startswith('#'):  # Skip empty lines and comments
                                continue
                            queries_list.append({'query': line, 'index': idx + 1})
                        
                        if not queries_list:
                            st.error("❌ Không tìm thấy lệnh SQL hợp lệ!")
                        else:
                            st.info(f"📊 Tìm thấy {len(queries_list)} lệnh SQL. Bắt đầu thực thi...")
                            
                            overall_progress = st.progress(0)
                            overall_status = st.empty()
                            
                            # Execute all queries and store results
                            all_results = []
                            for idx, query_info in enumerate(queries_list):
                                overall_status.text(f"⏳ Đang thực thi lệnh {idx + 1}/{len(queries_list)}")
                                
                                results = execute_query(st.session_state['db_connection'], query_info['query'])
                                
                                if results:
                                    all_results.append({
                                        'query': query_info['query'],
                                        'results': results,
                                        'row_count': len(results),
                                        'index': query_info['index']
                                    })
                                    st.success(f"✅ Lệnh {idx + 1}: {len(results)} dòng")
                                else:
                                    st.warning(f"⚠️ Lệnh {idx + 1}: Không có kết quả")
                                
                                overall_progress.progress((idx + 1) / len(queries_list))
                            
                            if not all_results:
                                st.error("❌ Không có kết quả nào từ các lệnh SQL!")
                            else:
                                # Store results in session state
                                st.session_state['batch_results'] = all_results
                                overall_status.text("✅ Hoàn thành thực thi!")
                                st.success(f"🎉 Đã thực thi thành công {len(all_results)} lệnh!")
                                
                    except Exception as e:
                        st.error(f"❌ Lỗi trong quá trình thực thi: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
        
        with col_exec2:
            if st.button("🗑️ Xóa kết quả", key="clear_batch"):
                st.session_state['batch_results'] = None
                st.rerun()
        
        # Step 2: Display results and allow file naming
        if st.session_state['batch_results']:
            st.markdown("---")
            st.markdown("### Bước 2: Xem kết quả và Cấu hình từng bảng")
            
            # Store configurations for each table
            table_configs = []
            
            # Display each result table with configuration options
            for idx, result_info in enumerate(st.session_state['batch_results']):
                with st.expander(f"📊 Bảng {idx + 1} - Lệnh SQL #{result_info['index']} ({result_info['row_count']} dòng)", expanded=True):
                    # Display query and preview
                    st.code(result_info['query'], language='sql')
                    st.dataframe(pd.DataFrame(result_info['results']).head(10), use_container_width=True)
                    if result_info['row_count'] > 10:
                        st.caption(f"Hiển thị 10/{result_info['row_count']} dòng đầu tiên")
                    
                    # Configuration for this table
                    st.markdown("**⚙️ Cấu hình xuất cho bảng này:**")
                    config_col1, config_col2 = st.columns(2)
                    
                    with config_col1:
                        default_name = f"query_{result_info['index']}"
                        file_name = st.text_input(
                            "📝 Tên file",
                            value=default_name,
                            key=f"filename_{idx}",
                            help="Tên file xuất (không cần phần mở rộng)"
                        )
                        
                        rows_per_file = st.number_input(
                            "📄 Số dòng mỗi file",
                            min_value=1,
                            value=min(50000, result_info['row_count']),
                            key=f"rows_{idx}",
                            help="Nếu bảng có nhiều dòng hơn, sẽ tự động chia thành nhiều file"
                        )
                    
                    with config_col2:
                        include_headers = st.checkbox(
                            "📋 Bao gồm tiêu đề cột",
                            value=True,
                            key=f"headers_{idx}"
                        )
                        
                        double_row = st.checkbox(
                            "📑 Xuất dòng kép (mỗi dòng lặp 2 lần)",
                            value=False,
                            key=f"double_{idx}"
                        )
                    
                    # Store config for this table
                    table_configs.append({
                        'file_name': file_name,
                        'rows_per_file': rows_per_file,
                        'include_headers': include_headers,
                        'double_row': double_row
                    })
            
            # Global export options
            st.markdown("---")
            st.markdown("### Bước 3: Tùy chọn xuất chung")
            
            col1, col2 = st.columns(2)
            with col1:
                batch_file_prefix = st.text_input("Prefix cho file ZIP", "batch_export", key="batch_prefix")
            with col2:
                batch_export_format = st.selectbox("Định dạng xuất", ["Excel (.xlsx)", "TXT (.txt)", "SQL (.sql)"], key="batch_format")
            
            # SQL table name for SQL export
            batch_sql_table = ""
            if batch_export_format == "SQL (.sql)":
                batch_sql_table = st.text_input("Tên bảng SQL (dùng chung cho tất cả)", key="batch_sql_table")
            
            # Export button
            if st.button("📥 Xuất tất cả file", key="export_batch", type="primary"):
                # Validate configurations
                file_names = [cfg['file_name'] for cfg in table_configs]
                
                if not all(file_names):
                    st.error("❌ Vui lòng đặt tên cho tất cả các file!")
                elif len(file_names) != len(set(file_names)):
                    st.error("❌ Tên file không được trùng lặp!")
                elif batch_export_format == "SQL (.sql)" and not batch_sql_table:
                    st.error("❌ Vui lòng nhập tên bảng SQL!")
                else:
                    try:
                        temp_dir = tempfile.mkdtemp()
                        overall_status = st.empty()
                        
                        # Export based on format
                        overall_status.text(f"📦 Đang tạo file xuất...")
                        zip_path = os.path.join(temp_dir, f"{batch_file_prefix}_batch.zip")
                        
                        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                            export_progress = st.progress(0)
                            export_status = st.empty()
                            
                            for idx, result_info in enumerate(st.session_state['batch_results']):
                                # Get configuration for this table
                                config = table_configs[idx]
                                file_name = config['file_name']
                                rows_per_file = config['rows_per_file']
                                include_headers = config['include_headers']
                                double_row = config['double_row']
                                results = result_info['results']
                                
                                export_status.text(f"📝 Đang xuất file {idx + 1}/{len(st.session_state['batch_results'])}: {file_name}")
                                
                                if batch_export_format == "Excel (.xlsx)":
                                    # Export to Excel
                                    total_rows = len(results)
                                    chunks = [results[i:i + rows_per_file] 
                                            for i in range(0, total_rows, rows_per_file)]
                                    
                                    for chunk_idx, chunk in enumerate(chunks):
                                        if len(chunks) > 1:
                                            excel_file_name = f'{file_name}-{chunk_idx+1:03d}.xlsx'
                                        else:
                                            excel_file_name = f'{file_name}.xlsx'
                                        
                                        file_path = os.path.join(temp_dir, excel_file_name)
                                        
                                        wb = openpyxl.Workbook()
                                        ws = wb.active
                                        headers = list(chunk[0].keys())
                                        row_num = 1
                                        
                                        if include_headers:
                                            for col, header in enumerate(headers, 1):
                                                cell = ws.cell(row=1, column=col, value=header)
                                                cell.font = Font(bold=True)
                                                cell.alignment = Alignment(horizontal='center')
                                            row_num = 2
                                        
                                        for record in chunk:
                                            for col, key in enumerate(headers, 1):
                                                ws.cell(row=row_num, column=col, value=str(record[key]))
                                            row_num += 1
                                            
                                            # Double row if enabled
                                            if double_row:
                                                for col, key in enumerate(headers, 1):
                                                    ws.cell(row=row_num, column=col, value=str(record[key]))
                                                row_num += 1
                                        
                                        wb.save(file_path)
                                        zipf.write(file_path, excel_file_name)
                                
                                elif batch_export_format == "TXT (.txt)":
                                    # Export to TXT
                                    total_rows = len(results)
                                    chunks = [results[i:i + rows_per_file] 
                                            for i in range(0, total_rows, rows_per_file)]
                                    
                                    for chunk_idx, chunk in enumerate(chunks):
                                        if len(chunks) > 1:
                                            txt_file_name = f'{file_name}-{chunk_idx+1:03d}.txt'
                                        else:
                                            txt_file_name = f'{file_name}.txt'
                                        
                                        file_path = os.path.join(temp_dir, txt_file_name)
                                        
                                        with open(file_path, 'w', encoding='utf-8', newline='') as txtfile:
                                            headers = list(chunk[0].keys())
                                            
                                            if include_headers:
                                                txtfile.write(','.join(headers) + '\n')
                                            
                                            for record in chunk:
                                                row_data = []
                                                for key in headers:
                                                    value = record[key]
                                                    if value is None:
                                                        row_data.append('')
                                                    else:
                                                        str_value = str(value).replace('\n', ' ').replace('\r', ' ')
                                                        if ',' in str_value:
                                                            str_value = f'"{str_value}"'
                                                        row_data.append(str_value)
                                                
                                                data_line = ','.join(row_data)
                                                txtfile.write(data_line + '\n')
                                                
                                                # Double row if enabled
                                                if double_row:
                                                    txtfile.write(data_line + '\n')
                                        
                                        zipf.write(file_path, txt_file_name)
                                
                                elif batch_export_format == "SQL (.sql)":
                                    # Export to SQL
                                    if not batch_sql_table:
                                        st.error("❌ Vui lòng nhập tên bảng SQL!")
                                        break
                                    
                                    total_rows = len(results)
                                    chunks = [results[i:i + rows_per_file] 
                                            for i in range(0, total_rows, rows_per_file)]
                                    
                                    for chunk_idx, chunk in enumerate(chunks):
                                        if len(chunks) > 1:
                                            sql_file_name = f'{file_name}-{chunk_idx+1:03d}.sql'
                                        else:
                                            sql_file_name = f'{file_name}.sql'
                                        
                                        file_path = os.path.join(temp_dir, sql_file_name)
                                        
                                        columns = list(chunk[0].keys())
                                        columns_str = ', '.join(f"`{col}`" for col in columns)
                                        
                                        values_list = []
                                        for record in chunk:
                                            values = []
                                            for col in columns:
                                                val = record[col]
                                                if val is None:
                                                    values.append('NULL')
                                                elif isinstance(val, (int, float)):
                                                    values.append(str(val))
                                                else:
                                                    val = str(val).replace('\\', '\\\\').replace("'", "\\'").replace('"', '\\"')
                                                    values.append(f"'{val}'")
                                            
                                            value_str = f"({', '.join(values)})"
                                            values_list.append(value_str)
                                            
                                            # Double row if enabled
                                            if double_row:
                                                values_list.append(value_str)
                                        
                                        insert_query = f"INSERT INTO `{batch_sql_table}` ({columns_str}) VALUES\n"
                                        insert_query += ',\n'.join(values_list) + ";\n"
                                        
                                        with open(file_path, 'w', encoding='utf-8') as f:
                                            f.write(insert_query)
                                        
                                        zipf.write(file_path, sql_file_name)
                                
                                export_progress.progress((idx + 1) / len(st.session_state['batch_results']))
                        
                        # Provide download button
                        with open(zip_path, 'rb') as f:
                            zip_contents = f.read()
                            st.download_button(
                                "📥 Tải xuất tất cả các file (ZIP)",
                                zip_contents,
                                file_name=f"{batch_file_prefix}_batch.zip",
                                mime="application/zip",
                                key="download_batch"
                            )
                        
                        # Summary
                        st.success(f"✅ Hoàn thành! Đã xuất {len(st.session_state['batch_results'])} file")
                        with st.expander("📊 Chi tiết kết quả"):
                            for idx, result_info in enumerate(st.session_state['batch_results']):
                                config = table_configs[idx]
                                st.write(f"**{config['file_name']}**: {result_info['row_count']} dòng")
                                st.caption(f"   ├─ Rows per file: {config['rows_per_file']} | Headers: {'✓' if config['include_headers'] else '✗'} | Double row: {'✓' if config['double_row'] else '✗'}")
                        
                        # Clean up
                        import shutil
                        shutil.rmtree(temp_dir, ignore_errors=True)
                        
                    except Exception as e:
                        st.error(f"❌ Lỗi trong quá trình xuất: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
    
    # Tab 6: Excel Upload & QR Code Generation
    with tab6:
        st.subheader("📤 Upload File Excel và Tạo QR Code")
        
        st.info("💡 Upload file Excel, chọn cột dữ liệu để tạo QR code và cột để đặt tên cho file ảnh QR")
        
        # File uploader for Excel
        uploaded_excel = st.file_uploader(
            "Chọn file Excel",
            type=['xlsx', 'xls'],
            key="excel_upload_qr",
            help="Upload file Excel chứa dữ liệu để tạo QR code"
        )
        
        if uploaded_excel is not None:
            try:
                # Read Excel file
                df_excel = pd.read_excel(uploaded_excel)
                
                st.success(f"✅ Đã tải file thành công! Tìm thấy {len(df_excel)} dòng và {len(df_excel.columns)} cột")
                
                # Display preview
                with st.expander("👁️ Xem trước dữ liệu", expanded=True):
                    st.dataframe(df_excel.head(20), use_container_width=True)
                    if len(df_excel) > 20:
                        st.caption(f"Hiển thị 20/{len(df_excel)} dòng đầu tiên")
                
                # Configuration section
                st.markdown("---")
                st.markdown("### ⚙️ Cấu hình tạo QR Code")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Select column for QR code data
                    qr_data_column = st.selectbox(
                        "📊 Chọn cột chứa dữ liệu QR Code",
                        options=df_excel.columns.tolist(),
                        key="excel_qr_column",
                        help="Dữ liệu từ cột này sẽ được mã hóa thành QR code"
                    )
                    
                    # Select QR format
                    qr_format_excel = st.selectbox(
                        "🖼️ Định dạng QR Code",
                        options=["png", "jpg", "svg"],
                        key="excel_qr_format"
                    )
                    
                    # QR Code size
                    qr_box_size = st.slider(
                        "📏 Kích thước QR Code",
                        min_value=5,
                        max_value=20,
                        value=10,
                        key="qr_box_size",
                        help="Kích thước ô vuông trong QR code"
                    )
                
                with col2:
                    # Select column for image filename
                    filename_column = st.selectbox(
                        "📝 Chọn cột để đặt tên file",
                        options=df_excel.columns.tolist(),
                        key="excel_filename_column",
                        help="Tên file ảnh QR sẽ được lấy từ cột này"
                    )
                    
                    # File prefix
                    excel_qr_prefix = st.text_input(
                        "📁 Prefix cho file ZIP",
                        value="qr_codes_from_excel",
                        key="excel_qr_prefix"
                    )
                    
                    # QR border
                    qr_border = st.slider(
                        "🔲 Viền QR Code",
                        min_value=1,
                        max_value=10,
                        value=5,
                        key="qr_border",
                        help="Độ rộng viền xung quanh QR code"
                    )
                
                # Additional options
                st.markdown("### 🔧 Tùy chọn nâng cao")
                
                col3, col4 = st.columns(2)
                with col3:
                    skip_empty = st.checkbox(
                        "⏭️ Bỏ qua dòng có giá trị trống",
                        value=True,
                        key="skip_empty",
                        help="Không tạo QR code cho các dòng có dữ liệu trống"
                    )
                
                with col4:
                    add_index_to_filename = st.checkbox(
                        "🔢 Thêm số thứ tự vào tên file",
                        value=False,
                        key="add_index",
                        help="Thêm số thứ tự vào đầu tên file để tránh trùng lặp"
                    )
                
                # Preview selected data
                st.markdown("### 👀 Xem trước dữ liệu sẽ tạo QR")
                preview_df = df_excel[[qr_data_column, filename_column]].head(10)
                preview_df.columns = ['Dữ liệu QR', 'Tên File']
                st.dataframe(preview_df, use_container_width=True)
                
                # Generate QR Codes button
                st.markdown("---")
                if st.button("🎨 Tạo QR Code", key="generate_qr_from_excel", type="primary"):
                    try:
                        temp_dir = tempfile.mkdtemp()
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        # Create zip file
                        zip_path = os.path.join(temp_dir, f"{excel_qr_prefix}.zip")
                        
                        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                            total_rows = len(df_excel)
                            success_count = 0
                            skip_count = 0
                            error_count = 0
                            
                            for idx, row in df_excel.iterrows():
                                try:
                                    # Get QR data
                                    qr_data = row[qr_data_column]
                                    
                                    # Skip if empty and skip_empty is enabled
                                    if skip_empty and (pd.isna(qr_data) or str(qr_data).strip() == ''):
                                        skip_count += 1
                                        continue
                                    
                                    # Get filename
                                    raw_filename = row[filename_column]
                                    
                                    # Skip if filename is empty
                                    if pd.isna(raw_filename) or str(raw_filename).strip() == '':
                                        if skip_empty:
                                            skip_count += 1
                                            continue
                                        else:
                                            raw_filename = f"qr_code_{idx + 1}"
                                    
                                    # Clean filename
                                    clean_name = str(raw_filename).replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                                    invalid_chars = '<>:"/\\|?*'
                                    for char in invalid_chars:
                                        clean_name = clean_name.replace(char, '_')
                                    clean_name = clean_name.strip('. ')
                                    
                                    # Limit filename length
                                    if len(clean_name) > 200:
                                        clean_name = clean_name[:200]
                                    
                                    # Ensure filename is not empty
                                    if not clean_name:
                                        clean_name = f"qr_code_{idx + 1}"
                                    
                                    # Add index if enabled
                                    if add_index_to_filename:
                                        clean_name = f"{idx + 1:05d}_{clean_name}"
                                    
                                    # Create full filename
                                    image_name = f"{clean_name}.{qr_format_excel}"
                                    file_path = os.path.join(temp_dir, image_name)
                                    
                                    # Generate QR code
                                    qr = qrcode.QRCode(
                                        version=1,
                                        box_size=qr_box_size,
                                        border=qr_border
                                    )
                                    qr.add_data(str(qr_data))
                                    qr.make(fit=True)
                                    
                                    # Save based on format
                                    if qr_format_excel == 'svg':
                                        factory = qrcode.image.svg.SvgPathImage
                                        qr_svg = qrcode.QRCode(
                                            version=1,
                                            box_size=qr_box_size,
                                            border=qr_border,
                                            image_factory=factory
                                        )
                                        qr_svg.add_data(str(qr_data))
                                        qr_svg.make(fit=True)
                                        img_svg = qr_svg.make_image(fill_color="black", back_color="white")
                                        img_svg.save(file_path)
                                    elif qr_format_excel == 'png':
                                        img = qr.make_image(fill_color="black", back_color="white")
                                        img.save(file_path, format='PNG')
                                    elif qr_format_excel == 'jpg':
                                        img = qr.make_image(fill_color="black", back_color="white")
                                        img_rgb = img.convert('RGB')
                                        img_rgb.save(file_path, format='JPEG', quality=95)
                                    
                                    # Add to zip
                                    zipf.write(file_path, image_name)
                                    success_count += 1
                                    
                                except Exception as e:
                                    error_count += 1
                                    st.warning(f"⚠️ Lỗi tại dòng {idx + 1}: {str(e)}")
                                
                                # Update progress
                                progress = (idx + 1) / total_rows
                                progress_bar.progress(progress)
                                status_text.text(f"⏳ Đang xử lý: {idx + 1}/{total_rows} | Thành công: {success_count} | Bỏ qua: {skip_count} | Lỗi: {error_count}")
                        
                        # Provide download button
                        if success_count > 0:
                            with open(zip_path, 'rb') as f:
                                zip_contents = f.read()
                                st.download_button(
                                    "📥 Tải xuống tất cả QR Code (ZIP)",
                                    zip_contents,
                                    file_name=f"{excel_qr_prefix}.zip",
                                    mime="application/zip",
                                    key="download_excel_qr"
                                )
                            
                            # Summary
                            st.success(f"✅ Hoàn thành! Đã tạo {success_count} QR code")
                            
                            with st.expander("📊 Thống kê chi tiết"):
                                st.write(f"**Tổng số dòng:** {total_rows}")
                                st.write(f"**✅ Thành công:** {success_count}")
                                st.write(f"**⏭️ Bỏ qua:** {skip_count}")
                                st.write(f"**❌ Lỗi:** {error_count}")
                        else:
                            st.error("❌ Không tạo được QR code nào. Vui lòng kiểm tra lại dữ liệu!")
                        
                        # Clean up
                        import shutil
                        shutil.rmtree(temp_dir, ignore_errors=True)
                        
                    except Exception as e:
                        st.error(f"❌ Lỗi trong quá trình tạo QR code: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
                
            except Exception as e:
                st.error(f"❌ Lỗi khi đọc file Excel: {str(e)}")
                st.info("💡 Đảm bảo file Excel của bạn có định dạng đúng (.xlsx hoặc .xls)")
    
    # Tab 7: Tra cứu mã lỗi từ file Excel
    with tab7:
        st.subheader("🔍 Tra cứu dữ liệu từ Serial/QR Code")
        
        st.info("💡 Upload file Excel chứa cột serial hoặc QR code. Hệ thống sẽ tự động phân biệt (số thuần túy = serial, sẽ tự động thêm prefix '26.' khi tra cứu) và truy vấn từ bảng codes_evnhcm")
        
        # Initialize session state for lookup results
        if 'lookup_results' not in st.session_state:
            st.session_state['lookup_results'] = None
        
        # File uploader
        uploaded_lookup_file = st.file_uploader(
            "📂 Chọn file Excel chứa Serial/QR Code",
            type=['xlsx', 'xls'],
            key="lookup_excel_upload",
            help="File Excel cần có ít nhất 1 cột chứa dữ liệu serial hoặc qrcode"
        )
        
        if uploaded_lookup_file is not None:
            try:
                # Read Excel file
                df_lookup = pd.read_excel(uploaded_lookup_file)
                
                st.success(f"✅ Đã tải file thành công! Tìm thấy {len(df_lookup)} dòng và {len(df_lookup.columns)} cột")
                
                # Display preview
                with st.expander("👁️ Xem trước dữ liệu", expanded=True):
                    st.dataframe(df_lookup.head(20), use_container_width=True)
                    if len(df_lookup) > 20:
                        st.caption(f"Hiển thị 20/{len(df_lookup)} dòng đầu tiên")
                
                # Configuration
                st.markdown("---")
                st.markdown("### ⚙️ Cấu hình tra cứu")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Select column containing serial/qrcode
                    data_column = st.selectbox(
                        "📊 Chọn cột chứa Serial/QR Code",
                        options=df_lookup.columns.tolist(),
                        key="lookup_data_column",
                        help="Cột chứa dữ liệu serial hoặc qrcode cần tra cứu"
                    )
                
                with col2:
                    # Additional columns to retrieve
                    st.markdown("**📋 Cột sẽ lấy từ database:**")
                    st.write("✓ qrcode")
                    st.write("✓ serial")
                    st.caption("Các cột mặc định từ bảng codes_evnhcm")
                
                # Preview selected data
                st.markdown("### 👀 Xem trước dữ liệu sẽ tra cứu")
                preview_lookup = df_lookup[[data_column]].head(10).copy()
                
                # Add column to show data type detection
                def detect_type(value):
                    if pd.isna(value):
                        return "❓ Trống"
                    str_value = str(value).strip()
                    if str_value.isdigit():
                        return "🔢 Serial"
                    else:
                        return "🔗 QR Code"
                
                def get_db_search_value(value):
                    if pd.isna(value):
                        return ""
                    str_value = str(value).strip()
                    if str_value.isdigit():
                        return f"26.{str_value}"
                    else:
                        return str_value
                
                preview_lookup['Loại dữ liệu'] = df_lookup[data_column].head(10).apply(detect_type)
                preview_lookup['Giá trị tìm trong DB'] = df_lookup[data_column].head(10).apply(get_db_search_value)
                preview_lookup.columns = ['Dữ liệu gốc', 'Loại phát hiện', 'Tra cứu trong DB']
                st.dataframe(preview_lookup, use_container_width=True)
                st.caption("💡 Serial sẽ tự động thêm prefix '26.' khi tra cứu trong database")
                
                # Lookup button
                st.markdown("---")
                if st.button("🔎 Bắt đầu tra cứu", key="start_lookup", type="primary"):
                    if not st.session_state['db_connection']:
                        st.error("❌ Vui lòng kết nối database trước!")
                    else:
                        try:
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            # Prepare data for lookup
                            lookup_data = df_lookup[data_column].tolist()
                            total_items = len(lookup_data)
                            
                            # Separate serials and qrcodes
                            serials = []
                            qrcodes = []
                            data_map = {}  # Map to store original order
                            
                            status_text.text("🔄 Đang phân loại dữ liệu...")
                            
                            for idx, value in enumerate(lookup_data):
                                if pd.isna(value):
                                    data_map[idx] = {'type': 'empty', 'value': value}
                                    continue
                                
                                str_value = str(value).strip()
                                if str_value.isdigit():
                                    # Thêm prefix "26." vào trước serial để tìm trong database
                                    serial_with_prefix = f"26.{str_value}"
                                    serials.append(serial_with_prefix)
                                    data_map[idx] = {'type': 'serial', 'value': str_value, 'db_value': serial_with_prefix}
                                else:
                                    qrcodes.append(str_value)
                                    data_map[idx] = {'type': 'qrcode', 'value': str_value}
                            
                            st.info(f"📊 Phân loại: {len(serials)} serial, {len(qrcodes)} qrcode, {total_items - len(serials) - len(qrcodes)} trống")
                            
                            # Query database
                            results_dict = {}
                            
                            # Query for serials
                            if serials:
                                status_text.text(f"🔍 Đang tra cứu {len(serials)} serial...")
                                serials_str = "', '".join(serials)
                                query_serial = f"SELECT `qrcode`, `serial` FROM codes_evnhcm WHERE `serial` IN ('{serials_str}')"
                                
                                serial_results = execute_query(st.session_state['db_connection'], query_serial)
                                
                                if serial_results:
                                    for result in serial_results:
                                        results_dict[result['serial']] = result
                                    st.success(f"✅ Tìm thấy {len(serial_results)}/{len(serials)} serial trong database")
                                else:
                                    st.warning(f"⚠️ Không tìm thấy serial nào trong database")
                            
                            progress_bar.progress(0.5)
                            
                            # Query for qrcodes
                            if qrcodes:
                                status_text.text(f"🔍 Đang tra cứu {len(qrcodes)} qrcode...")
                                qrcodes_str = "', '".join(qrcodes)
                                query_qrcode = f"SELECT `qrcode`, `serial` FROM codes_evnhcm WHERE `qrcode` IN ('{qrcodes_str}')"
                                
                                qrcode_results = execute_query(st.session_state['db_connection'], query_qrcode)
                                
                                if qrcode_results:
                                    for result in qrcode_results:
                                        results_dict[result['qrcode']] = result
                                    st.success(f"✅ Tìm thấy {len(qrcode_results)}/{len(qrcodes)} qrcode trong database")
                                else:
                                    st.warning(f"⚠️ Không tìm thấy qrcode nào trong database")
                            
                            progress_bar.progress(0.8)
                            
                            # Build results in original order
                            status_text.text("📝 Đang sắp xếp kết quả...")
                            
                            ordered_results = []
                            found_count = 0
                            not_found_count = 0
                            
                            for idx in range(total_items):
                                item = data_map.get(idx, {})
                                item_type = item.get('type')
                                item_value = item.get('value')
                                
                                if item_type == 'empty':
                                    ordered_results.append({
                                        'STT': idx + 1,
                                        'Dữ liệu gốc': item_value,
                                        'Loại': '❓ Trống',
                                        'qrcode': '',
                                        'serial': '',
                                        'Trạng thái': '⚠️ Trống'
                                    })
                                    not_found_count += 1
                                elif item_type == 'serial':
                                    # Sử dụng db_value (có prefix 26.) để tìm trong results_dict
                                    db_value = item.get('db_value', item_value)
                                    result = results_dict.get(db_value)
                                    if result:
                                        ordered_results.append({
                                            'STT': idx + 1,
                                            'Dữ liệu gốc': item_value,
                                            'Loại': '🔢 Serial',
                                            'qrcode': result.get('qrcode', ''),
                                            'serial': result.get('serial', ''),
                                            'Trạng thái': '✅ Tìm thấy'
                                        })
                                        found_count += 1
                                    else:
                                        ordered_results.append({
                                            'STT': idx + 1,
                                            'Dữ liệu gốc': item_value,
                                            'Loại': '🔢 Serial',
                                            'qrcode': '',
                                            'serial': '',
                                            'Trạng thái': '❌ Không tìm thấy'
                                        })
                                        not_found_count += 1
                                elif item_type == 'qrcode':
                                    result = results_dict.get(item_value)
                                    if result:
                                        ordered_results.append({
                                            'STT': idx + 1,
                                            'Dữ liệu gốc': item_value,
                                            'Loại': '🔗 QR Code',
                                            'qrcode': result.get('qrcode', ''),
                                            'serial': result.get('serial', ''),
                                            'Trạng thái': '✅ Tìm thấy'
                                        })
                                        found_count += 1
                                    else:
                                        ordered_results.append({
                                            'STT': idx + 1,
                                            'Dữ liệu gốc': item_value,
                                            'Loại': '🔗 QR Code',
                                            'qrcode': '',
                                            'serial': '',
                                            'Trạng thái': '❌ Không tìm thấy'
                                        })
                                        not_found_count += 1
                            
                            progress_bar.progress(1.0)
                            status_text.text("✅ Hoàn thành tra cứu!")
                            
                            # Store results in session state
                            st.session_state['lookup_results'] = ordered_results
                            
                            # Display summary
                            st.success(f"🎉 Hoàn thành! Tìm thấy: {found_count}/{total_items} | Không tìm thấy: {not_found_count}/{total_items}")
                            
                        except Exception as e:
                            st.error(f"❌ Lỗi trong quá trình tra cứu: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
                
            except Exception as e:
                st.error(f"❌ Lỗi khi đọc file Excel: {str(e)}")
                st.info("💡 Đảm bảo file Excel của bạn có định dạng đúng (.xlsx hoặc .xls)")
        
        # Display results if available
        if st.session_state['lookup_results']:
            st.markdown("---")
            st.markdown("### 📊 Kết quả tra cứu")
            
            df_results = pd.DataFrame(st.session_state['lookup_results'])
            
            # Display statistics
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            with col_stat1:
                total = len(df_results)
                st.metric("📝 Tổng số dòng", total)
            with col_stat2:
                found = len(df_results[df_results['Trạng thái'] == '✅ Tìm thấy'])
                st.metric("✅ Tìm thấy", found)
            with col_stat3:
                not_found = len(df_results[df_results['Trạng thái'].isin(['❌ Không tìm thấy', '⚠️ Trống'])])
                st.metric("❌ Không tìm thấy", not_found)
            
            # Display results table
            st.dataframe(df_results, use_container_width=True, height=400)
            
            # Export options
            st.markdown("---")
            st.markdown("### 📥 Xuất kết quả")
            
            col_export1, col_export2, col_export3, col_export4 = st.columns(4)
            
            with col_export1:
                # Export to Excel
                if st.button("📊 Xuất Excel đầy đủ", key="export_lookup_excel"):
                    try:
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_results.to_excel(writer, index=False, sheet_name='Kết quả tra cứu')
                        
                        output.seek(0)
                        st.download_button(
                            "⬇️ Tải xuống file Excel",
                            output.getvalue(),
                            file_name="ket_qua_tra_cuu.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_lookup_excel"
                        )
                        st.success("✅ File Excel đã sẵn sàng!")
                    except Exception as e:
                        st.error(f"❌ Lỗi khi tạo file Excel: {str(e)}")
            
            with col_export2:
                # Export to CSV
                if st.button("📄 Xuất CSV", key="export_lookup_csv"):
                    try:
                        csv_data = df_results.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            "⬇️ Tải xuống file CSV",
                            csv_data,
                            file_name="ket_qua_tra_cuu.csv",
                            mime="text/csv",
                            key="download_lookup_csv"
                        )
                        st.success("✅ File CSV đã sẵn sàng!")
                    except Exception as e:
                        st.error(f"❌ Lỗi khi tạo file CSV: {str(e)}")
            
            with col_export3:
                # Export only found results
                if st.button("✅ Xuất kết quả tìm thấy", key="export_found_only"):
                    try:
                        df_found = df_results[df_results['Trạng thái'] == '✅ Tìm thấy']
                        
                        if len(df_found) == 0:
                            st.warning("⚠️ Không có kết quả nào được tìm thấy!")
                        else:
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                df_found.to_excel(writer, index=False, sheet_name='Tìm thấy')
                            
                            output.seek(0)
                            st.download_button(
                                "⬇️ Tải xuống (chỉ tìm thấy)",
                                output.getvalue(),
                                file_name="ket_qua_tim_thay.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_found_excel"
                            )
                            st.success(f"✅ File Excel với {len(df_found)} kết quả đã sẵn sàng!")
                    except Exception as e:
                        st.error(f"❌ Lỗi khi tạo file: {str(e)}")
            
            with col_export4:
                # Export 2 columns: qrcode with URL prefix and serial
                if st.button("🔗 Xuất QR+Serial", key="export_qr_serial"):
                    try:
                        # Filter only found results
                        df_found = df_results[df_results['Trạng thái'] == '✅ Tìm thấy']
                        
                        if len(df_found) == 0:
                            st.warning("⚠️ Không có kết quả nào được tìm thấy!")
                        else:
                            # Create new dataframe with 2 columns
                            df_export = pd.DataFrame()
                            
                            # Add URL prefix to qrcode
                            df_export['qrcode'] = df_found['qrcode'].apply(
                                lambda x: f"http://evnhcm.vinachg.com/?s={x}" if pd.notna(x) and str(x).strip() != '' else ''
                            )
                            df_export['serial'] = df_found['serial']
                            
                            # Create Excel file
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                df_export.to_excel(writer, index=False, sheet_name='QR và Serial')
                                
                                # Format the worksheet
                                worksheet = writer.sheets['QR và Serial']
                                
                                # Set column widths
                                worksheet.column_dimensions['A'].width = 50
                                worksheet.column_dimensions['B'].width = 20
                                
                                # Style headers
                                for cell in worksheet[1]:
                                    cell.font = Font(bold=True)
                                    cell.alignment = Alignment(horizontal='center')
                            
                            output.seek(0)
                            st.download_button(
                                "⬇️ Tải xuống QR+Serial",
                                output.getvalue(),
                                file_name="qrcode_serial.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_qr_serial"
                            )
                            st.success(f"✅ File Excel với {len(df_export)} dòng đã sẵn sàng!")
                            st.info("📋 File chứa 2 cột: qrcode (có URL đầy đủ) và serial")
                    except Exception as e:
                        st.error(f"❌ Lỗi khi tạo file: {str(e)}")

if __name__ == "__main__":
    main()
